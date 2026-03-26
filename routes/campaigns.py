"""
Routes Campagnes
"""
import os, uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from extensions import db
from models import User, Campaign, AccountReview
import pandas as pd
from openpyxl import load_workbook

campaigns_bp = Blueprint('campaigns', __name__)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')

def get_current_user():
    uid = get_jwt_identity()
    return User.query.get(int(uid)) if uid else None

PROFIL_DESCRIPTIONS = {
    'administrateur': 'Accès complet en écriture et administration — risque élevé',
    'admin':          'Accès complet en écriture et administration — risque élevé',
    'contributeur':   'Accès en lecture et écriture sur les données métier',
    'lecteur':        'Accès en lecture seule — risque faible',
    'lecture':        'Accès en lecture seule — risque faible',
    'n/a':            'Compte présent dans l\'AD sans profil applicatif défini',
}

def get_profil_description(profil):
    if not profil: return ''
    return PROFIL_DESCRIPTIONS.get(profil.lower().strip(), f'Profil {profil} — vérifier les droits associés')

@campaigns_bp.route('', methods=['GET'])
@jwt_required()
def list_campaigns():
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403
    camps = Campaign.query.order_by(Campaign.created_at.desc()).all()
    return jsonify([c.to_dict() for c in camps])

@campaigns_bp.route('/assigned', methods=['GET'])
@jwt_required()
def assigned_campaigns():
    current = get_current_user()
    if not current: return jsonify({'error': 'Non authentifié'}), 401
    # Cherche par email AD ET par email profil mis à jour
    emails_to_check = [current.email]
    if current.email_ad and current.email_ad != current.email:
        emails_to_check.append(current.email_ad)

    camp_ids = db.session.query(AccountReview.campaign_id)\
        .filter(AccountReview.manager_email.in_(emails_to_check))\
        .distinct().all()
    camp_ids = [c[0] for c in camp_ids]
    camps = Campaign.query.filter(Campaign.id.in_(camp_ids)).all()

    result = []
    for c in camps:
        my_accounts = AccountReview.query.filter(
            AccountReview.campaign_id == c.id,
            AccountReview.manager_email.in_(emails_to_check)
        ).all()
        result.append({
            **c.to_dict(),
            'total':      len(my_accounts),
            'decided':    sum(1 for a in my_accounts if a.decision),
            'pending':    sum(1 for a in my_accounts if not a.decision),
            'risk_count': sum(1 for a in my_accounts if a.score > 0),
        })
    return jsonify(result)

@campaigns_bp.route('/<int:camp_id>', methods=['GET'])
@jwt_required()
def get_campaign(camp_id):
    current = get_current_user()
    if not current: return jsonify({'error': 'Non authentifié'}), 401
    c = Campaign.query.get_or_404(camp_id)
    return jsonify(c.to_dict())

@campaigns_bp.route('/<int:camp_id>', methods=['PUT'])
@jwt_required()
def update_campaign(camp_id):
    """Modifier une campagne (nom, description, échéance) ou re-uploader les fichiers."""
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403

    camp = Campaign.query.get_or_404(camp_id)

    # Mise à jour des champs texte
    if request.is_json:
        data = request.json or {}
        if 'nom'         in data: camp.nom         = data['nom']
        if 'description' in data: camp.description = data['description']
        if 'dateEcheance'in data and data['dateEcheance']:
            camp.date_echeance = datetime.strptime(data['dateEcheance'], '%Y-%m-%d')
        db.session.commit()
        return jsonify({'ok': True, 'campaign': camp.to_dict()})

    # Re-analyse avec nouveaux fichiers
    if 'fileApp' in request.files or 'fileAD' in request.files:
        if 'fileApp' not in request.files or 'fileAD' not in request.files:
            return jsonify({'error': 'Les deux fichiers sont requis pour re-analyser'}), 400

        inactivity_days  = int(request.form.get('inactivityDays', camp.inactivity_days))
        sensitive_raw    = request.form.get('sensitiveGroups', camp.sensitive_groups)
        sensitive_groups = [g.strip() for g in sensitive_raw.split(',') if g.strip()]

        file_app = request.files['fileApp']
        file_ad  = request.files['fileAD']
        sid      = str(uuid.uuid4())[:8]
        path_app = os.path.join(UPLOAD_FOLDER, f'{sid}_app.xlsx')
        path_ad  = os.path.join(UPLOAD_FOLDER, f'{sid}_ad.xlsx')
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file_app.save(path_app)
        file_ad.save(path_ad)

        try:
            # Supprimer anciens comptes
            AccountReview.query.filter_by(campaign_id=camp.id).delete()

            from iam_access_review import (
                load_excel, normalize_dates, check_required_columns,
                detect_orphans, detect_inactive, detect_privileged, build_report,
                style_report, add_stats_sheet
            )
            df_ad  = load_excel(path_ad,  'AD')
            df_app = load_excel(path_app, 'Application')
            check_required_columns(df_ad,  ['sAMAccountName', 'memberOf', 'Last Logon Date'], 'AD')
            check_required_columns(df_app, ['Account_ID'], 'Application')
            df_ad  = normalize_dates(df_ad,  ['Last Logon Date', 'D entrée société', 'D sortie société'])
            df_app = normalize_dates(df_app, ['Date création compte', 'Dernière connexion app'])

            orphans_app, orphans_ad = detect_orphans(df_ad, df_app)
            inactive_mask           = detect_inactive(df_ad, inactivity_days)
            priv_mask, priv_groups  = detect_privileged(df_ad, sensitive_groups)
            df_report               = build_report(df_ad, df_app, orphans_app, orphans_ad,
                                                   inactive_mask, priv_mask, priv_groups, inactivity_days)

            camp.inactivity_days  = inactivity_days
            camp.sensitive_groups = sensitive_raw

            inactif_col = f'Inactif (>{inactivity_days}j)'
            for _, row in df_report.iterrows():
                profil = str(row.get('Profil App', '') or '')
                acc = AccountReview(
                    campaign_id        = camp.id,
                    account_id         = str(row.get('Account_ID', '')),
                    user_id            = str(row.get('User_ID', '')),
                    nom_complet        = str(row.get('Nom complet', '')),
                    profil_app         = profil,
                    profil_description = get_profil_description(profil),
                    statut_app         = str(row.get('Statut App', '')),
                    direction          = str(row.get('Direction', '')),
                    nature_contact     = str(row.get('Nature contact', '')),
                    job_title          = str(row.get('Job Title', '')),
                    manager_matricule  = str(row.get('Matricule manager', '')),
                    manager_nom        = str(row.get('Nom manager', '')),
                    manager_email      = str(row.get('E-mail manager', '')),
                    last_logon_ad      = str(row.get('Last Logon AD', '')),
                    groupes_sensibles  = str(row.get('Groupes sensibles', '')),
                    d_sortie           = str(row.get('D sortie société', '')),
                    orphelin_app       = str(row.get('Orphelin (App)', 'NON')) == 'OUI',
                    orphelin_ad        = str(row.get('Orphelin (AD)',  'NON')) == 'OUI',
                    inactif            = str(row.get(inactif_col, 'NON')) == 'OUI',
                    privilegie         = str(row.get('Compte privilégié', 'NON')) == 'OUI',
                    score              = int(row.get('Score risque', 0)),
                    libelle_risque     = str(row.get('Libellé risque', '')),
                )
                db.session.add(acc)

            report_path = os.path.join(UPLOAD_FOLDER, f'camp_{camp.id}_rapport.xlsx')
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df_report.to_excel(writer, index=False, sheet_name='Revue des accès')
            wb = load_workbook(report_path)
            style_report(wb['Revue des accès'], df_report, inactivity_days)
            add_stats_sheet(wb, df_report, inactivity_days)
            wb.save(report_path)

            db.session.commit()
            return jsonify({'ok': True, 'campaign': camp.to_dict(),
                            'stats': {'total': len(df_report)}}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            for p in [path_app, path_ad]:
                if os.path.exists(p): os.remove(p)

    return jsonify({'error': 'Aucune donnée à mettre à jour'}), 400


@campaigns_bp.route('/<int:camp_id>', methods=['DELETE'])
@jwt_required()
def delete_campaign(camp_id):
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403
    camp = Campaign.query.get_or_404(camp_id)
    db.session.delete(camp)
    db.session.commit()
    return jsonify({'ok': True})


@campaigns_bp.route('/<int:camp_id>/archive', methods=['POST'])
@jwt_required()
def archive_campaign(camp_id):
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403
    camp = Campaign.query.get_or_404(camp_id)
    camp.statut = 'archivee' if camp.statut != 'archivee' else 'en_cours'
    db.session.commit()
    return jsonify({'ok': True, 'statut': camp.statut})


@campaigns_bp.route('/<int:camp_id>/accounts', methods=['GET'])
@jwt_required()
def get_accounts(camp_id):
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403
    accounts = AccountReview.query.filter_by(campaign_id=camp_id)\
               .order_by(AccountReview.score.desc()).all()
    return jsonify([a.to_dict() for a in accounts])


@campaigns_bp.route('/<int:camp_id>/accounts/mine', methods=['GET'])
@jwt_required()
def get_my_accounts(camp_id):
    current = get_current_user()
    if not current: return jsonify({'error': 'Non authentifié'}), 401
    emails_to_check = [current.email]
    if current.email_ad and current.email_ad != current.email:
        emails_to_check.append(current.email_ad)
    accounts = AccountReview.query.filter(
        AccountReview.campaign_id == camp_id,
        AccountReview.manager_email.in_(emails_to_check)
    ).order_by(AccountReview.score.desc()).all()
    return jsonify([a.to_dict() for a in accounts])


@campaigns_bp.route('/<int:camp_id>/accounts/<int:acc_id>/decision', methods=['POST'])
@jwt_required()
def save_decision(camp_id, acc_id):
    current = get_current_user()
    if not current: return jsonify({'error': 'Non authentifié'}), 401
    data     = request.json or {}
    decision = data.get('decision', '')
    if decision and decision not in ('Maintenir', 'Révoquer', 'Investiguer'):
        return jsonify({'error': 'Décision invalide'}), 400
    acc = AccountReview.query.filter_by(id=acc_id, campaign_id=camp_id).first_or_404()
    acc.decision    = decision or None
    acc.decision_by = current.id
    acc.decision_at = datetime.utcnow() if decision else None
    db.session.commit()
    return jsonify({'ok': True, 'decision': acc.decision})


@campaigns_bp.route('', methods=['POST'])
@jwt_required()
def create_campaign():
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403

    nom             = request.form.get('nom', '').strip()
    description     = request.form.get('description', '').strip()
    inactivity_days = int(request.form.get('inactivityDays', 120))
    sensitive_raw   = request.form.get('sensitiveGroups', 'Domain Admins,Enterprise Admins')
    sensitive_groups= [g.strip() for g in sensitive_raw.split(',') if g.strip()]
    date_echeance   = request.form.get('dateEcheance')

    if not nom: return jsonify({'error': 'Nom de campagne requis'}), 400
    if 'fileApp' not in request.files or 'fileAD' not in request.files:
        return jsonify({'error': 'Les deux fichiers sont requis'}), 400

    file_app = request.files['fileApp']
    file_ad  = request.files['fileAD']
    sid      = str(uuid.uuid4())[:8]
    path_app = os.path.join(UPLOAD_FOLDER, f'{sid}_app.xlsx')
    path_ad  = os.path.join(UPLOAD_FOLDER, f'{sid}_ad.xlsx')
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    file_app.save(path_app)
    file_ad.save(path_ad)

    try:
        from iam_access_review import (
            load_excel, normalize_dates, check_required_columns,
            detect_orphans, detect_inactive, detect_privileged, build_report,
            style_report, add_stats_sheet
        )
        df_ad  = load_excel(path_ad,  'AD')
        df_app = load_excel(path_app, 'Application')
        check_required_columns(df_ad,  ['sAMAccountName', 'memberOf', 'Last Logon Date'], 'AD')
        check_required_columns(df_app, ['Account_ID'], 'Application')
        df_ad  = normalize_dates(df_ad,  ['Last Logon Date', 'D entrée société', 'D sortie société'])
        df_app = normalize_dates(df_app, ['Date création compte', 'Dernière connexion app'])

        orphans_app, orphans_ad = detect_orphans(df_ad, df_app)
        inactive_mask           = detect_inactive(df_ad, inactivity_days)
        priv_mask, priv_groups  = detect_privileged(df_ad, sensitive_groups)
        df_report               = build_report(df_ad, df_app, orphans_app, orphans_ad,
                                               inactive_mask, priv_mask, priv_groups, inactivity_days)

        camp = Campaign(
            nom=nom, description=description,
            inactivity_days=inactivity_days, sensitive_groups=sensitive_raw,
            created_by=current.id,
            date_echeance=datetime.strptime(date_echeance, '%Y-%m-%d') if date_echeance else None,
        )
        db.session.add(camp)
        db.session.flush()

        inactif_col = f'Inactif (>{inactivity_days}j)'
        for _, row in df_report.iterrows():
            profil = str(row.get('Profil App', '') or '')
            acc = AccountReview(
                campaign_id=camp.id,
                account_id=str(row.get('Account_ID', '')),
                user_id=str(row.get('User_ID', '')),
                nom_complet=str(row.get('Nom complet', '')),
                profil_app=profil,
                profil_description=get_profil_description(profil),
                statut_app=str(row.get('Statut App', '')),
                direction=str(row.get('Direction', '')),
                nature_contact=str(row.get('Nature contact', '')),
                job_title=str(row.get('Job Title', '')),
                manager_matricule=str(row.get('Matricule manager', '')),
                manager_nom=str(row.get('Nom manager', '')),
                manager_email=str(row.get('E-mail manager', '')),
                last_logon_ad=str(row.get('Last Logon AD', '')),
                groupes_sensibles=str(row.get('Groupes sensibles', '')),
                d_sortie=str(row.get('D sortie société', '')),
                orphelin_app=str(row.get('Orphelin (App)', 'NON')) == 'OUI',
                orphelin_ad=str(row.get('Orphelin (AD)', 'NON')) == 'OUI',
                inactif=str(row.get(inactif_col, 'NON')) == 'OUI',
                privilegie=str(row.get('Compte privilégié', 'NON')) == 'OUI',
                score=int(row.get('Score risque', 0)),
                libelle_risque=str(row.get('Libellé risque', '')),
            )
            db.session.add(acc)

        report_path = os.path.join(UPLOAD_FOLDER, f'camp_{camp.id}_rapport.xlsx')
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            df_report.to_excel(writer, index=False, sheet_name='Revue des accès')
        wb = load_workbook(report_path)
        style_report(wb['Revue des accès'], df_report, inactivity_days)
        add_stats_sheet(wb, df_report, inactivity_days)
        wb.save(report_path)

        db.session.commit()
        return jsonify({
            'campaign': camp.to_dict(),
            'stats': {
                'total': len(df_report),
                'orphan': int(orphans_app.sum() + orphans_ad.sum()),
                'inactive': int(inactive_mask.sum()),
                'privileged': int(priv_mask.sum()),
            }
        }), 201

    except SystemExit:
        db.session.rollback()
        return jsonify({'error': 'Colonnes obligatoires manquantes dans les fichiers'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        for p in [path_app, path_ad]:
            if os.path.exists(p): os.remove(p)


@campaigns_bp.route('/<int:camp_id>/export', methods=['GET'])
@jwt_required()
def export_campaign(camp_id):
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403
    report_path = os.path.join(UPLOAD_FOLDER, f'camp_{camp_id}_rapport.xlsx')
    if not os.path.exists(report_path):
        return jsonify({'error': 'Rapport non disponible'}), 404
    camp = Campaign.query.get_or_404(camp_id)
    return send_file(report_path, as_attachment=True,
                     download_name=f'rapport_{camp.nom.replace(" ","_")}_{camp.date_lancement.strftime("%Y%m%d")}.xlsx')


@campaigns_bp.route('/<int:camp_id>/dashboard', methods=['GET'])
@jwt_required()
def campaign_dashboard(camp_id):
    current = get_current_user()
    if not current or current.role != 'admin':
        return jsonify({'error': 'Accès admin requis'}), 403

    accounts = AccountReview.query.filter_by(campaign_id=camp_id).all()

    global_stats = {
        'total':      len(accounts),
        'orphan_app': sum(1 for a in accounts if a.orphelin_app),
        'orphan_ad':  sum(1 for a in accounts if a.orphelin_ad),
        'inactive':   sum(1 for a in accounts if a.inactif),
        'privileged': sum(1 for a in accounts if a.privilegie),
        'decided':    sum(1 for a in accounts if a.decision),
        'maintenir':  sum(1 for a in accounts if a.decision == 'Maintenir'),
        'revoquer':   sum(1 for a in accounts if a.decision == 'Révoquer'),
        'investiguer':sum(1 for a in accounts if a.decision == 'Investiguer'),
    }

    dirs = {}
    for a in accounts:
        d = a.direction or 'Non défini'
        if d not in dirs:
            dirs[d] = {'direction': d, 'total': 0, 'risk': 0, 'decided': 0, 'maintenir': 0, 'revoquer': 0}
        dirs[d]['total']     += 1
        dirs[d]['risk']      += 1 if a.score > 0 else 0
        dirs[d]['decided']   += 1 if a.decision else 0
        dirs[d]['maintenir'] += 1 if a.decision == 'Maintenir' else 0
        dirs[d]['revoquer']  += 1 if a.decision == 'Révoquer'  else 0

    managers = {}
    for a in accounts:
        m = a.manager_nom or 'Orphelin — sans manager'
        e = a.manager_email or ''
        if m not in managers:
            managers[m] = {'manager': m, 'email': e, 'total': 0, 'risk': 0, 'decided': 0}
        managers[m]['total']   += 1
        managers[m]['risk']    += 1 if a.score > 0 else 0
        managers[m]['decided'] += 1 if a.decision else 0
        # Mise à jour email si on a mieux
        if e and not managers[m]['email']:
            managers[m]['email'] = e

    return jsonify({
        'global':       global_stats,
        'by_direction': sorted(dirs.values(),     key=lambda x: x['total'],   reverse=True),
        'by_manager':   sorted(managers.values(), key=lambda x: x['decided'], reverse=True),
    })
