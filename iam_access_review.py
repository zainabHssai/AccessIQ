"""
=============================================================
  IAM Access Review — Script MVP
  Basé sur le Cahier des Charges v1.0
=============================================================
  Fonctionnalités :
    1. Import des fichiers Excel (AD + Application)
    2. Normalisation des colonnes et des dates
    3. Matching (clé : sAMAccountName ↔ Account_ID)
    4. Détection des comptes à risque :
         - Orphelins (non mappés)
         - Inactifs (seuil configurable)
         - Privilèges excessifs (groupes sensibles configurables)
    5. Génération d'un rapport Excel enrichi
    6. Statistiques de synthèse

  Usage :
    python iam_access_review.py \
        --ad  ad_extract.xlsx \
        --app app_extract.xlsx \
        [--inactivity-days 120] \
        [--sensitive-groups "Domain Admins,Enterprise Admins"] \
        [--output rapport_revue.xlsx]
=============================================================
"""

import argparse
import sys
import os
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────
# PARAMÈTRES PAR DÉFAUT (modifiables sans changer le code)
# ─────────────────────────────────────────────────────────────
DEFAULT_INACTIVITY_DAYS = 120          # seuil inactivité
DEFAULT_SENSITIVE_GROUPS = [           # groupes à surveiller
    "Domain Admins",
    "Enterprise Admins",
    "Schema Admins",
    "Backup Operators",
]
DEFAULT_OUTPUT = "rapport_revue_acces.xlsx"

# Mapping des variantes de noms de colonnes → noms standards
COLUMN_ALIASES = {
    # AD
    "samaccountname": "sAMAccountName",
    "sam account name": "sAMAccountName",
    "login": "sAMAccountName",
    "account": "sAMAccountName",
    "lastlogon": "Last Logon Date",
    "last logon": "Last Logon Date",
    "lastlogondate": "Last Logon Date",
    "last logon date": "Last Logon Date",
    "lastlogontimestamp": "Last Logon Date",
    "memberof": "memberOf",
    "member of": "memberOf",
    "groupes": "memberOf",
    "groups": "memberOf",
    # App
    "account_id": "Account_ID",
    "accountid": "Account_ID",
    "identifiant": "Account_ID",
    "user_id": "User_ID",
    "userid": "User_ID",
    "matricule": "User_ID",
    "profil": "Profil",
    "role": "Profil",
    "rôle": "Profil",
    "dernière connexion app": "Dernière connexion app",
    "derniere connexion app": "Dernière connexion app",
    "last logon app": "Dernière connexion app",
}

REQUIRED_AD_COLS  = ["sAMAccountName", "memberOf", "Last Logon Date"]
REQUIRED_APP_COLS = ["Account_ID"]

# Couleurs rapport
COLOR = {
    "header_ad":        "1F3864",  # bleu foncé
    "header_app":       "14375E",
    "header_risk":      "7B0000",
    "header_action":    "375623",
    "orphan":           "F4CCCC",  # rouge/rose — Orphelin (App sans AD) — risque élevé
    "non_provisionne":  "D9E8F5",  # bleu clair — Non Provisionné (AD sans App) — risque faible
    "inactive":         "FFF2CC",  # jaune — inactif
    "privileged":       "E2EFDA",  # vert clair — privilégié
    "multi_risk":       "F9B8B8",  # rose foncé — plusieurs risques
    "normal":           "FFFFFF",
    "alt_row":          "F5F5F5",
}


# ═══════════════════════════════════════════════════════════════
# 1. CHARGEMENT ET NORMALISATION
# ═══════════════════════════════════════════════════════════════

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Renomme les colonnes selon le mapping d'alias."""
    new_cols = {}
    for col in df.columns:
        key = col.strip().lower()
        if key in COLUMN_ALIASES:
            new_cols[col] = COLUMN_ALIASES[key]
    return df.rename(columns=new_cols)


def parse_date(val) -> datetime | None:
    """Convertit une valeur en datetime (gère 'Never', NaN, formats variés)."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.lower() in ("never", "jamais", "n/a", "", "0", "none"):
        return None
    # Formats courants
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S",
                "%d-%m-%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt)
        except ValueError:
            pass
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None


def normalize_dates(df: pd.DataFrame, date_cols: list) -> pd.DataFrame:
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(parse_date)
    return df


def check_required_columns(df: pd.DataFrame, required: list, source: str):
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"\n❌ ERREUR — Colonnes obligatoires manquantes dans {source} :")
        for m in missing:
            print(f"   • {m}")
        print("\n   Colonnes présentes :", list(df.columns))
        sys.exit(1)


def load_excel(path: str, label: str) -> pd.DataFrame:
    if not os.path.exists(path):
        print(f"❌ Fichier introuvable : {path}")
        sys.exit(1)
    df = pd.read_excel(path)
    df = normalize_columns(df)
    print(f"✅ {label} chargé — {len(df)} lignes, colonnes : {list(df.columns)}")
    return df


# ═══════════════════════════════════════════════════════════════
# 2. DÉTECTION DES RISQUES
# ═══════════════════════════════════════════════════════════════

def detect_orphans(df_ad: pd.DataFrame, df_app: pd.DataFrame):
    """
    Retourne :
      orphans_app : comptes APP sans correspondance dans AD
      orphans_ad  : comptes AD sans correspondance dans APP
    """
    ad_accounts  = set(df_ad["sAMAccountName"].dropna().str.lower())
    app_accounts = set(df_app["Account_ID"].dropna().str.lower())

    orphans_app_mask = ~df_app["Account_ID"].str.lower().isin(ad_accounts)
    orphans_ad_mask  = ~df_ad["sAMAccountName"].str.lower().isin(app_accounts)

    return orphans_app_mask, orphans_ad_mask


def detect_inactive(df_ad: pd.DataFrame, threshold_days: int) -> pd.Series:
    """Retourne un masque booléen des comptes inactifs dans AD."""
    cutoff = datetime.now() - timedelta(days=threshold_days)

    def is_inactive(val):
        if val is None:
            return True   # jamais connecté = inactif
        return val < cutoff

    return df_ad["Last Logon Date"].apply(is_inactive)


def detect_privileged(df_ad: pd.DataFrame, sensitive_groups: list) -> pd.Series:
    """Retourne un masque booléen + la colonne des groupes sensibles détectés."""
    sg_lower = [g.lower() for g in sensitive_groups]

    def find_sensitive(member_of_val):
        if pd.isna(member_of_val) or str(member_of_val).strip() == "":
            return ""
        groups = [g.strip() for g in str(member_of_val).split(";")]
        found = [g for g in groups if g.lower() in sg_lower]
        return ";".join(found)

    sensitive_col = df_ad["memberOf"].apply(find_sensitive)
    mask = sensitive_col != ""
    return mask, sensitive_col


# ═══════════════════════════════════════════════════════════════
# 3. CONSTRUCTION DU RAPPORT CONSOLIDÉ
# ═══════════════════════════════════════════════════════════════

def build_report(df_ad, df_app, orphans_app_mask, orphans_ad_mask,
                 inactive_mask, priv_mask, priv_groups,
                 threshold_days) -> pd.DataFrame:
    """
    Construit un DataFrame consolidé avec toutes les informations
    et les indicateurs de risque.
    """
    today_str = datetime.now().strftime("%Y-%m-%d")

    # ── Construire un dictionnaire Matricule → Email depuis l'AD
    # Cherche les colonnes même si leur nom varie légèrement
    def find_col(df, candidates):
        """Trouve une colonne par nom parmi plusieurs variantes (insensible à la casse)."""
        cols_lower = {c.lower().strip(): c for c in df.columns}
        for candidate in candidates:
            if candidate.lower() in cols_lower:
                return cols_lower[candidate.lower()]
        return None

    col_matricule = find_col(df_ad, [
        "Matricule", "matricule", "MATRICULE", "mat", "MAT",
        "ID", "id", "employee_id", "employeeid", "Identifiant",
        "User_ID", "user_id", "userid"
    ])
    col_email = find_col(df_ad, [
        "E-mail professionnel", "email professionnel", "Email professionnel",
        "E-mail", "email", "Email", "mail", "Mail",
        "E-mail pro", "EmailPro", "email_pro",
        "userPrincipalName", "upn"
    ])
    col_manager_mat = find_col(df_ad, [
        "Matricule manager (N+2)", "matricule manager (n+2)",
        "Matricule manager", "matricule manager",
        "Manager matricule", "manager_matricule",
        "ManagerID", "managerid", "Manager ID",
        "Matricule N+2", "matricule n+2"
    ])

    matricule_to_email = {}
    if col_matricule and col_email:
        for _, row in df_ad.iterrows():
            mat   = str(row.get(col_matricule, "")).strip()
            email = str(row.get(col_email, "")).strip()
            if mat and email and mat.lower() not in ("nan", "", "none") \
                    and email.lower() not in ("nan", "", "none"):
                matricule_to_email[mat] = email

    # Log pour debug
    print(f"  [Email manager] Colonnes détectées :")
    print(f"    Matricule      : '{col_matricule}'")
    print(f"    Email          : '{col_email}'")
    print(f"    Manager mat    : '{col_manager_mat}'")
    print(f"    Dict construit : {len(matricule_to_email)} entrées")

    def get_manager_email(ad_row):
        """Résout l'email du manager via son matricule dans l'AD."""
        if not col_manager_mat:
            return ""
        manager_mat = str(ad_row.get(col_manager_mat, "")).strip()
        if manager_mat and manager_mat.lower() not in ("nan", "", "none"):
            return matricule_to_email.get(manager_mat, "")
        return ""

    # --- Partie 1 : comptes APP matchés ou orphelins côté APP ---
    rows = []
    for _, app_row in df_app.iterrows():
        acc_id = str(app_row["Account_ID"]).lower().strip()
        # Cherche correspondance dans AD
        ad_match = df_ad[df_ad["sAMAccountName"].str.lower() == acc_id]
        is_orphan_app = len(ad_match) == 0

        if is_orphan_app:
            row = {
                # Infos App
                "Account_ID":           app_row.get("Account_ID", ""),
                "User_ID":              app_row.get("User_ID", ""),
                "Nom complet":          app_row.get("Nom complet", ""),
                "Profil App":           app_row.get("Profil", ""),
                "Statut App":           app_row.get("Statut compte", ""),
                "Date création compte": str(app_row.get("Date création compte", "")),
                "Dern. connexion App":  str(app_row.get("Dernière connexion app", "")),
                # Infos AD (vides)
                "Matricule AD":         "",
                "Nom AD":               "",
                "Prénom AD":            "",
                "Direction":            "",
                "Nature contact":       "",
                "Job Title":            "",
                "Matricule manager":    "",
                "Nom manager":          "",
                "E-mail manager":       "",
                "Last Logon AD":        "",
                "Groupes sensibles":    "",
                "D sortie société":     "",
                # Risques
                "Orphelin":             "OUI",
                "Non Provisionné":      "NON",
                f"Inactif (>{threshold_days}j)": "N/A",
                "Compte privilégié":    "NON",
                "Score risque":         2,
                "Libellé risque":       "Orphelin",
                # Décision manager
                "Décision manager":     "",
                "Motif":                "",
                "Date analyse":         today_str,
            }
        else:
            ad = ad_match.iloc[0]
            last_logon = ad.get("Last Logon Date")
            is_inactive = (last_logon is None) or \
                          (last_logon < datetime.now() - timedelta(days=threshold_days))
            ad_acc = ad.get("sAMAccountName", "")
            ad_idx = df_ad[df_ad["sAMAccountName"] == ad_acc].index
            is_priv = bool(priv_mask.loc[ad_idx].any()) if len(ad_idx) else False
            grp_sensibles = priv_groups.loc[ad_idx].values[0] if len(ad_idx) else ""

            risks = []
            if is_inactive:  risks.append("Inactif")
            if is_priv:      risks.append("Privilégié")
            score = len(risks)

            row = {
                "Account_ID":           app_row.get("Account_ID", ""),
                "User_ID":              app_row.get("User_ID", ""),
                "Nom complet":          app_row.get("Nom complet", ""),
                "Profil App":           app_row.get("Profil", ""),
                "Statut App":           app_row.get("Statut compte", ""),
                "Date création compte": str(app_row.get("Date création compte", "")),
                "Dern. connexion App":  str(app_row.get("Dernière connexion app", "")),
                "Matricule AD":         ad.get("Matricule", ""),
                "Nom AD":               ad.get("Nom", ""),
                "Prénom AD":            ad.get("Prénom", ""),
                "Direction":            ad.get("Direction", ""),
                "Nature contact":       ad.get("Nature contact", ""),
                "Job Title":            ad.get("Job Title", ""),
                "Matricule manager":    ad.get("Matricule manager (N+2)", ""),
                "Nom manager":          ad.get("Nom manager (N+2)", ""),
                "E-mail manager":       get_manager_email(ad),
                "Last Logon AD":        str(last_logon.date() if last_logon else "Jamais"),
                "Groupes sensibles":    grp_sensibles,
                "D sortie société":     str(ad.get("D sortie société", "")),
                "Orphelin":             "NON",
                "Non Provisionné":      "NON",
                f"Inactif (>{threshold_days}j)": "OUI" if is_inactive else "NON",
                "Compte privilégié":    "OUI" if is_priv else "NON",
                "Score risque":         score,
                "Libellé risque":       ";".join(risks) if risks else "Aucun",
                "Décision manager":     "",
                "Motif":                "",
                "Date analyse":         today_str,
            }
        rows.append(row)

    # --- Partie 2 : comptes AD orphelins (dans AD mais pas dans App) ---
    for idx, ad_row in df_ad[orphans_ad_mask].iterrows():
        last_logon = ad_row.get("Last Logon Date")
        is_inactive = (last_logon is None) or \
                      (last_logon < datetime.now() - timedelta(days=threshold_days))
        is_priv = bool(priv_mask.loc[idx]) if idx in priv_mask.index else False
        grp_sensibles = priv_groups.loc[idx] if idx in priv_groups.index else ""

        risks = []
        if is_inactive:  risks.append("Inactif")
        if is_priv:      risks.append("Privilégié")
        # Non Provisionné = score 1 (risque faible, dans AD mais pas dans App)
        score = 1 + len(risks)

        row = {
            "Account_ID":           ad_row.get("sAMAccountName", ""),
            "User_ID":              ad_row.get("Matricule", ""),
            "Nom complet":          f"{ad_row.get('Nom','')} {ad_row.get('Prénom','')}".strip(),
            "Profil App":           "N/A",
            "Statut App":           "N/A",
            "Date création compte": "",
            "Dern. connexion App":  "",
            "Matricule AD":         ad_row.get("Matricule", ""),
            "Nom AD":               ad_row.get("Nom", ""),
            "Prénom AD":            ad_row.get("Prénom", ""),
            "Direction":            ad_row.get("Direction", ""),
            "Nature contact":       ad_row.get("Nature contact", ""),
            "Job Title":            ad_row.get("Job Title", ""),
            "Matricule manager":    ad_row.get("Matricule manager (N+2)", ""),
            "Nom manager":          ad_row.get("Nom manager (N+2)", ""),
            "E-mail manager":       get_manager_email(ad_row),
            "Last Logon AD":        str(last_logon.date() if last_logon else "Jamais"),
            "Groupes sensibles":    grp_sensibles,
            "D sortie société":     str(ad_row.get("D sortie société", "")),
            "Orphelin":             "NON",
            "Non Provisionné":      "OUI",
            f"Inactif (>{threshold_days}j)": "OUI" if is_inactive else "NON",
            "Compte privilégié":    "OUI" if is_priv else "NON",
            "Score risque":         score,
            "Libellé risque":       "Non Provisionné" + ((";" + ";".join(risks)) if risks else ""),
            "Décision manager":     "",
            "Motif":                "",
            "Date analyse":         today_str,
        }
        rows.append(row)

    df_report = pd.DataFrame(rows)
    # Trier par score risque décroissant
    df_report = df_report.sort_values("Score risque", ascending=False).reset_index(drop=True)
    return df_report


# ═══════════════════════════════════════════════════════════════
# 4. MISE EN FORME EXCEL
# ═══════════════════════════════════════════════════════════════

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_report(ws, df, threshold_days):
    inactif_col = f"Inactif (>{threshold_days}j)"

    # En-têtes
    header_fills = {
        "Account_ID": COLOR["header_app"],
        "User_ID":    COLOR["header_app"],
        "Nom complet":COLOR["header_app"],
        "Profil App": COLOR["header_app"],
        "Statut App": COLOR["header_app"],
        "Date création compte": COLOR["header_app"],
        "Dern. connexion App":  COLOR["header_app"],
        "Matricule AD":  COLOR["header_ad"],
        "Nom AD":        COLOR["header_ad"],
        "Prénom AD":     COLOR["header_ad"],
        "Direction":     COLOR["header_ad"],
        "Nature contact":COLOR["header_ad"],
        "Job Title":     COLOR["header_ad"],
        "Matricule manager": COLOR["header_ad"],
        "Nom manager":       COLOR["header_ad"],
        "E-mail manager":    COLOR["header_ad"],
        "Last Logon AD":     COLOR["header_ad"],
        "Groupes sensibles": COLOR["header_ad"],
        "D sortie société":  COLOR["header_ad"],
        "Orphelin":          COLOR["header_risk"],
        "Non Provisionné":   COLOR["header_risk"],
        inactif_col:         COLOR["header_risk"],
        "Compte privilégié": COLOR["header_risk"],
        "Score risque":      COLOR["header_risk"],
        "Libellé risque":    COLOR["header_risk"],
        "Décision manager":  COLOR["header_action"],
        "Motif":             COLOR["header_action"],
        "Date analyse":      COLOR["header_action"],
        "Date décision":     COLOR["header_action"],
    }

    col_names = list(df.columns)
    for c_idx, col_name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.value = col_name
        fill_hex = header_fills.get(col_name, "2F5597")
        cell.fill = make_fill(fill_hex)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()

    ws.row_dimensions[1].height = 35

    # Colonnes de risque pour la logique de couleur
    col_map = {name: idx + 1 for idx, name in enumerate(col_names)}

    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2
        is_orphan_app = str(row.get("Orphelin",        "NON")).upper() == "OUI"
        is_orphan_ad  = str(row.get("Non Provisionné", "NON")).upper() == "OUI"
        is_inactive   = str(row.get(inactif_col, "NON")).upper() == "OUI"
        is_priv       = str(row.get("Compte privilégié", "NON")).upper() == "OUI"
        score         = int(row.get("Score risque", 0))

        # Couleur de ligne
        if score >= 2 and is_orphan_app:
            row_fill = make_fill(COLOR["multi_risk"])
        elif is_orphan_app:
            row_fill = make_fill(COLOR["orphan"])
        elif is_orphan_ad and score >= 2:
            row_fill = make_fill(COLOR["inactive"])   # Non Provisionné + autre risque
        elif is_orphan_ad:
            row_fill = make_fill(COLOR["non_provisionne"])
        elif is_inactive and is_priv:
            row_fill = make_fill(COLOR["multi_risk"])
        elif is_inactive:
            row_fill = make_fill(COLOR["inactive"])
        elif is_priv:
            row_fill = make_fill(COLOR["privileged"])
        else:
            row_fill = make_fill(COLOR["alt_row"] if r_idx % 2 else COLOR["normal"])

        for c_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=excel_row, column=c_idx)
            val = row[col_name]
            cell.value = str(val) if val is not None else ""
            cell.fill = row_fill
            cell.border = make_border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(size=10)

    # Largeurs colonnes (auto)
    col_widths = {
        "Account_ID": 18, "User_ID": 12, "Nom complet": 25,
        "Profil App": 16, "Statut App": 14,
        "Date création compte": 20, "Dern. connexion App": 20,
        "Matricule AD": 14, "Nom AD": 18, "Prénom AD": 18,
        "Direction": 18, "Nature contact": 15, "Job Title": 22,
        "Matricule manager": 18, "Nom manager": 20, "E-mail manager": 28,
        "Last Logon AD": 18, "Groupes sensibles": 30, "D sortie société": 16,
        "Orphelin": 14, "Non Provisionné": 18, inactif_col: 18,
        "Compte privilégié": 18, "Score risque": 13, "Libellé risque": 25,
        "Décision manager": 20, "Motif": 28, "Date analyse": 14,
    }
    for col_name, width in col_widths.items():
        if col_name in col_map:
            ws.column_dimensions[get_column_letter(col_map[col_name])].width = width

    # Validation liste déroulante — Décision manager
    dec_col_letter = get_column_letter(col_map.get("Décision manager", 1))
    dv_decision = DataValidation(
        type="list",
        formula1='"Maintenu,Révoqué,Recertifié,À investiguer"',
        allow_blank=True,
        showDropDown=False
    )
    dv_decision.sqref = f"{dec_col_letter}2:{dec_col_letter}5000"
    ws.add_data_validation(dv_decision)

    # Validation liste déroulante — Motif
    motif_col_letter = get_column_letter(col_map.get("Motif", 1))
    dv_motif = DataValidation(
        type="list",
        formula1='"Compte métier actif,Départ définitif,Compte prestataire expiré,Compte de service,Accès légitime confirmé,Accès à restreindre,Autre"',
        allow_blank=True,
        showDropDown=False
    )
    dv_motif.sqref = f"{motif_col_letter}2:{motif_col_letter}5000"
    ws.add_data_validation(dv_motif)

    # Figer la première ligne
    ws.freeze_panes = "A2"


def add_stats_sheet(wb, df, threshold_days):
    ws = wb.create_sheet("Statistiques")
    inactif_col = f"Inactif (>{threshold_days}j)"

    total           = len(df)
    n_orphan_app    = (df["Orphelin"]        == "OUI").sum()
    n_orphan_ad     = (df["Non Provisionné"] == "OUI").sum()
    n_inactive      = (df[inactif_col] == "OUI").sum()
    n_priv          = (df["Compte privilégié"] == "OUI").sum()
    n_no_risk       = (df["Score risque"] == 0).sum()
    n_multi_risk    = (df["Score risque"] >= 2).sum()

    stats = [
        ("📊 Statistiques de la revue d'accès", ""),
        ("Date d'analyse", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Seuil inactivité (jours)", threshold_days),
        ("", ""),
        ("Indicateur", "Valeur"),
        ("Total comptes analysés", total),
        ("Comptes sans risque détecté", n_no_risk),
        ("Comptes orphelins (App sans AD)", n_orphan_app),
        ("Comptes non provisionnés (AD sans App)", n_orphan_ad),
        ("Comptes inactifs", n_inactive),
        ("Comptes à privilèges excessifs", n_priv),
        ("Comptes à risques multiples (score ≥ 2)", n_multi_risk),
        ("", ""),
        ("Répartition par direction", ""),
    ]

    for r_idx, (label, val) in enumerate(stats, 1):
        ws.cell(row=r_idx, column=1, value=label).font = Font(
            bold=(r_idx in (1, 5, 14)), size=11 if r_idx == 1 else 10
        )
        ws.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            ws.cell(row=r_idx, column=1).fill = make_fill("1F3864")
            ws.cell(row=r_idx, column=1).font = Font(bold=True, color="FFFFFF", size=12)
        if r_idx == 5:
            for c in (1, 2):
                ws.cell(row=r_idx, column=c).fill = make_fill("2F5597")
                ws.cell(row=r_idx, column=c).font = Font(bold=True, color="FFFFFF")

    # Répartition par direction (on remplace les directions vides par "Non identifié — orphelin App")
    df_stats = df.copy()
    df_stats["Direction"] = df_stats["Direction"].replace("", "Non identifié (orphelin App)")
    df_stats["Direction"] = df_stats["Direction"].fillna("Non identifié (orphelin App)")

    start = len(stats) + 1
    dir_counts = df_stats.groupby("Direction").size().reset_index(name="Nb comptes")
    risk_by_dir = df_stats[df_stats["Score risque"] > 0].groupby("Direction").size().reset_index(name="Nb risques")
    dir_stats = dir_counts.merge(risk_by_dir, on="Direction", how="left").fillna(0)

    ws.cell(row=start, column=1, value="Direction").fill = make_fill("2F5597")
    ws.cell(row=start, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start, column=2, value="Nb comptes").fill = make_fill("2F5597")
    ws.cell(row=start, column=2).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start, column=3, value="Nb comptes à risque").fill = make_fill("2F5597")
    ws.cell(row=start, column=3).font = Font(bold=True, color="FFFFFF")

    for i, row in dir_stats.iterrows():
        ws.cell(row=start + i + 1, column=1, value=row["Direction"])
        ws.cell(row=start + i + 1, column=2, value=int(row["Nb comptes"]))
        ws.cell(row=start + i + 1, column=3, value=int(row["Nb risques"]))

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22


# ═══════════════════════════════════════════════════════════════
# 5. POINT D'ENTRÉE
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="IAM Access Review — détection des comptes à risque"
    )
    parser.add_argument("--ad",  required=True,  help="Fichier Excel extract AD")
    parser.add_argument("--app", required=True,  help="Fichier Excel extract Application")
    parser.add_argument("--inactivity-days", type=int, default=DEFAULT_INACTIVITY_DAYS,
                        help=f"Seuil d'inactivité en jours (défaut : {DEFAULT_INACTIVITY_DAYS})")
    parser.add_argument("--sensitive-groups", type=str,
                        default=",".join(DEFAULT_SENSITIVE_GROUPS),
                        help="Groupes sensibles séparés par des virgules")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help=f"Nom du fichier rapport (défaut : {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    sensitive_groups = [g.strip() for g in args.sensitive_groups.split(",")]
    threshold_days   = args.inactivity_days

    print("\n" + "═" * 60)
    print("  IAM Access Review — Démarrage de l'analyse")
    print("═" * 60)
    print(f"  Seuil inactivité  : {threshold_days} jours")
    print(f"  Groupes sensibles : {sensitive_groups}")
    print("═" * 60 + "\n")

    # 1. Chargement
    df_ad  = load_excel(args.ad,  "Extract AD")
    df_app = load_excel(args.app, "Extract Application")

    # 2. Vérification colonnes obligatoires
    check_required_columns(df_ad,  REQUIRED_AD_COLS,  "AD")
    check_required_columns(df_app, REQUIRED_APP_COLS, "Application")

    # 3. Normalisation dates
    df_ad  = normalize_dates(df_ad,  ["Last Logon Date", "D entrée société", "D sortie société"])
    df_app = normalize_dates(df_app, ["Date création compte", "Dernière connexion app"])

    # 4. Détections
    orphans_app_mask, orphans_ad_mask = detect_orphans(df_ad, df_app)
    inactive_mask                     = detect_inactive(df_ad, threshold_days)
    priv_mask, priv_groups            = detect_privileged(df_ad, sensitive_groups)

    print(f"\n📋 Résultats :")
    print(f"   Comptes App sans correspondance AD : {orphans_app_mask.sum()}")
    print(f"   Comptes AD  sans correspondance App: {orphans_ad_mask.sum()}")
    print(f"   Comptes inactifs (AD)              : {inactive_mask.sum()}")
    print(f"   Comptes privilégiés (AD)           : {priv_mask.sum()}")

    # 5. Rapport consolidé
    df_report = build_report(
        df_ad, df_app,
        orphans_app_mask, orphans_ad_mask,
        inactive_mask, priv_mask, priv_groups,
        threshold_days
    )

    # 6. Export Excel mis en forme
    out_path = args.output
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_report.to_excel(writer, index=False, sheet_name="Revue des accès")

    wb = load_workbook(out_path)
    ws = wb["Revue des accès"]
    style_report(ws, df_report, threshold_days)
    add_stats_sheet(wb, df_report, threshold_days)
    wb.save(out_path)

    print(f"\n✅ Rapport généré : {out_path}")
    print(f"   {len(df_report)} comptes analysés au total\n")


if __name__ == "__main__":
    main()
